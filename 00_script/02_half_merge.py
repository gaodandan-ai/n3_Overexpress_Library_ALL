import openpyxl
import os
import re

def process_sheet_data(ws30, ws40, target_ws, numbers_range):
    """
    Helper function to process a specific range of numbers (e.g., 1-6 or 7-12)
    and write to the target worksheet.
    """
    letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
    
    # Get header rows for mapping
    header_30 = [cell.value for cell in ws30[1]]
    header_40 = [cell.value for cell in ws40[1]]
    
    # Pre-calculate column indices for each letter group
    letter_mapping = {}
    for char in letters:
        target_cols = [f"{char}{num}" for num in numbers_range]
        
        indices_30 = []
        for col_name in target_cols:
            try:
                indices_30.append(header_30.index(col_name))
            except ValueError:
                print(f"Warning: Column {col_name} not found in sheet {ws30.title}")
                
            indices_40 = []
            for col_name in target_cols:
                try:
                    indices_40.append(header_40.index(col_name))
                except ValueError:
                    print(f"Warning: Column {col_name} not found in sheet {ws40.title}")
        
        letter_mapping[char] = {
            '30': indices_30,
            '40': indices_40
        }
            
    # Process rows
    max_row = ws30.max_row
    
    for row_idx in range(1, max_row + 1):
        new_row = []
        
        # 1. Keep first two columns of ws30
        new_row.append(ws30.cell(row=row_idx, column=1).value)
        new_row.append(ws30.cell(row=row_idx, column=2).value)
        
        if row_idx == 1:
            # Special handling for header row: generate A1-A12, B1-B12, etc.
            for char in letters:
                count = len(letter_mapping[char]['30']) + len(letter_mapping[char]['40'])
                for i in range(1, count + 1):
                    new_row.append(f"{char}{i}")
        else:
            # 2. Add Letter+Numbers groups data iteratively
            for char in letters:
                # Add from ws30
                for col_idx in letter_mapping[char]['30']:
                    new_row.append(ws30.cell(row=row_idx, column=col_idx + 1).value)
                
                # Add from ws40
                for col_idx in letter_mapping[char]['40']:
                    new_row.append(ws40.cell(row=row_idx, column=col_idx + 1).value)
            
        target_ws.append(new_row)

def perform_merge_tasks(wb30, wb40, tasks, output_path):
    """
    Executes a list of specific merge tasks and saves to output_path.
    Each task is: (source_sheet_index, (r_start, r_end), target_title)
    """
    new_wb = openpyxl.Workbook()
    
    for i, (src_idx, (r_start, r_end), title) in enumerate(tasks):
        ws30 = wb30.worksheets[src_idx]
        ws40 = wb40.worksheets[src_idx]
        
        if i == 0:
            target_ws = new_wb.active
            target_ws.title = title
        else:
            target_ws = new_wb.create_sheet(title=title)
            
        print(f"  -> Processing sheet: {title}...")
        process_sheet_data(ws30, ws40, target_ws, [str(num) for num in range(r_start, r_end)])
            
    print(f"Saving to {output_path}...")
    new_wb.save(output_path)

def merge_excel():
    # Base directory is the parent of this script's directory
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    
    # Directories to process: (source_directory, output_directory)
    dirs_to_process = [
        (os.path.join(base_dir, r'04_result\OD\01_average_merge'), os.path.join(base_dir, r'04_result\OD\02_half_merge')),
        (os.path.join(base_dir, r'04_result\R\01_average_merge'), os.path.join(base_dir, r'04_result\R\02_half_merge'))
    ]
    
    for src_dir, out_dir in dirs_to_process:
        print(f"\n========================================\nProcessing directory: {src_dir}")
        if not os.path.exists(src_dir):
            print(f"Directory not found: {src_dir}")
            continue
            
        os.makedirs(out_dir, exist_ok=True)
        
        # Find all _30.xlsx files, skipping temporary Excel files
        files_30 = sorted([f for f in os.listdir(src_dir) if f.endswith('_30.xlsx') and not f.startswith('~$')])
        
        if not files_30:
            print(f"No _30.xlsx files found in {src_dir}")
            continue
            
        for f30 in files_30:
            f40 = f30.replace('_30.xlsx', '_40.xlsx')
            path_30 = os.path.join(src_dir, f30)
            path_40 = os.path.join(src_dir, f40)
            
            if not os.path.exists(path_40):
                print(f"Warning: Corresponding _40.xlsx file for {f30} not found at {path_40}. Skipping.")
                continue
                
            # Parse output names from prefix (e.g. Z1Z2 or Z1Z2_R)
            prefix = f30.replace('_30.xlsx', '')
            match_key = re.match(r'^Z(\d+)Z(\d+)(_R)?$', prefix)
            if not match_key:
                print(f"Warning: Filename prefix '{prefix}' does not match expected pattern ZxZy or ZxZy_R. Skipping.")
                continue
                
            z_start = match_key.group(1)       # e.g., "1"
            z_end = match_key.group(2)         # e.g., "2"
            suffix = match_key.group(3) or ""  # e.g., "_R" or ""
            
            out_file_1 = f"Z{z_start}{suffix}.xlsx"  # e.g. Z1.xlsx or Z1_R.xlsx
            out_file_2 = f"Z{z_end}{suffix}.xlsx"    # e.g. Z2.xlsx or Z2_R.xlsx
            
            out_path_1 = os.path.join(out_dir, out_file_1)
            out_path_2 = os.path.join(out_dir, out_file_2)
            
            print(f"\n>>> Merging {f30} and {f40} into {out_file_1} and {out_file_2}...")
            print(f"Loading {f30}...")
            wb30 = openpyxl.load_workbook(path_30, data_only=True)
            print(f"Loading {f40}...")
            wb40 = openpyxl.load_workbook(path_40, data_only=True)
            
            # Tasks configuration for the first file (e.g. Z1 / Z3 / Z5)
            # Source Sheet Indices: 
            # 0 -> 1(1-4), 1 -> 1(5-8), 2 -> 1(9-12), 6 -> H
            tasks_01 = [
                (0, (1, 7), "MTP1(1-2)"), (0, (7, 13), "MTP2(3-4)"),
                (1, (1, 7), "MTP3(5-6)"), (1, (7, 13), "MTP4(7-8)"),
                (2, (1, 7), "MTP5(9-10)"), (2, (7, 13), "MTP6(11-12)"),
                (6, (1, 7), "MTP7(last)")
            ]
            
            # Tasks configuration for the second file (e.g. Z2 / Z4 / Z6)
            # Source Sheet Indices: 
            # 3 -> 2(1-4), 4 -> 2(5-8), 5 -> 2(9-12), 6 -> H
            tasks_02 = [
                (3, (1, 7), "MTP1(1-2)"), (3, (7, 13), "MTP2(3-4)"),
                (4, (1, 7), "MTP3(5-6)"), (4, (7, 13), "MTP4(7-8)"),
                (5, (1, 7), "MTP5(9-10)"), (5, (7, 13), "MTP6(11-12)"),
                (6, (7, 13), "MTP7(last)")
            ]
            
            print(f"Generating {out_file_1}...")
            perform_merge_tasks(wb30, wb40, tasks_01, out_path_1)
            
            print(f"Generating {out_file_2}...")
            perform_merge_tasks(wb30, wb40, tasks_02, out_path_2)
            
    print("\n========================================\nAll tasks completed successfully!")

if __name__ == "__main__":
    merge_excel()
