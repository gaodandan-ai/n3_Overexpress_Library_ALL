import openpyxl
import shutil
import os

def swap_columns_r():
    file_path = r'F:\Growth_Profiler_Data\n3_Overexpress_Library_ALL\03_raw_data_r\Z1Z2_R_30.xlsx'
    backup_path = r'F:\Growth_Profiler_Data\n3_Overexpress_Library_ALL\03_raw_data_r\Z1Z2_R_30_original_backup.xlsx.bak'
    
    # 1. Create a backup for safety
    if not os.path.exists(backup_path):
        print(f"Creating safety backup at: {backup_path}")
        shutil.copy2(file_path, backup_path)
    else:
        print("Safety backup already exists.")
        
    # 2. Load the workbook
    print("Loading workbook...")
    wb = openpyxl.load_workbook(file_path)
    ws = wb['H']
    
    # 3. Get header columns and find indices (1-based)
    headers = [cell.value for cell in ws[1]]
    
    swaps = [
        # (col_name_1, col_name_2)
        ('A4', 'F4'), ('A5', 'F5'), ('A6', 'F6'),
        ('B4', 'E4'), ('B5', 'E5'), ('B6', 'E6'),
        ('C4', 'D4'), ('C5', 'D5'), ('C6', 'D6')
    ]
    
    swap_indices = []
    for c1, c2 in swaps:
        if c1 not in headers or c2 not in headers:
            print(f"Error: Could not find columns {c1} or {c2} in headers!")
            return
        idx1 = headers.index(c1) + 1  # openpyxl columns are 1-indexed
        idx2 = headers.index(c2) + 1
        swap_indices.append((c1, idx1, c2, idx2))
        
    print("\nColumns to swap:")
    for c1, idx1, c2, idx2 in swap_indices:
        print(f"  - {c1} (Col {idx1}) <-> {c2} (Col {idx2})")
        
    # 4. Print row 2 values before swap as verification
    print("\n--- Row 2 values BEFORE swap ---")
    for c1, idx1, c2, idx2 in swap_indices:
        val1 = ws.cell(row=2, column=idx1).value
        val2 = ws.cell(row=2, column=idx2).value
        print(f"  {c1}: {val1}  |  {c2}: {val2}")
        
    # 5. Perform the swaps for all data rows (row 2 to max_row)
    max_row = ws.max_row
    print(f"\nPerforming swaps on {max_row - 1} data rows...")
    for r in range(2, max_row + 1):
        for _, idx1, _, idx2 in swap_indices:
            temp = ws.cell(row=r, column=idx1).value
            ws.cell(row=r, column=idx1).value = ws.cell(row=r, column=idx2).value
            ws.cell(row=r, column=idx2).value = temp
            
    # 6. Print row 2 values after swap as verification
    print("\n--- Row 2 values AFTER swap ---")
    for c1, idx1, _, _ in swap_indices:
        val1 = ws.cell(row=2, column=idx1).value
        # corresponding column name 2
        c2 = [pair[2] for pair in swap_indices if pair[0] == c1][0]
        idx2 = [pair[3] for pair in swap_indices if pair[0] == c1][0]
        val2 = ws.cell(row=2, column=idx2).value
        print(f"  {c1}: {val1}  |  {c2}: {val2}")
        
    # 7. Save the workbook
    print(f"\nSaving modified workbook to: {file_path}")
    wb.save(file_path)
    print("Done successfully!")

if __name__ == '__main__':
    swap_columns_r()
